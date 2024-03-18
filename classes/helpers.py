import random
from openpyxl import load_workbook
from PyQt5.QtWidgets import QDialog
from .dahua import *
import traceback
from threading import Thread

wb = load_workbook("./classes/sitesNoDss.xlsx")
ws = wb.active


class Funcoes:
    def __init__(self, janela: QDialog) -> None:
        self._janela = janela

    @staticmethod
    def pegalistaip() -> dict:
        dicionario_de_pontos = {}
        for linha in ws.iter_rows(min_row=2):
            dicionario_de_pontos[linha[0].value] = linha[1].value
        return dicionario_de_pontos

    def coletar_seriais(self) -> None:
        """
        Esse método carrega os dados dos serias das câmeras e atualiza a progress bar e o textfield resultado,
        se selecionado o checkbox ele executa em todos os pontos da lista.
        """
        if not self._janela.qcb_todas_cameras.isChecked():
            self._janela.qpb_processo.setRange(0, len(self._janela.portas))
            self._janela.qte_resultado.clear()
            try:
                for progresso, porta in enumerate(self._janela.portas, 1):
                    ponto = self._janela.pontoComboBox.currentText()
                    retorno = gerar_lista_serial(self._janela.sites[ponto], porta)
                    if retorno:
                        self._janela.qte_resultado.appendPlainText(
                            f"{ponto};{porta};{retorno} \n"
                        )
                    self._janela.qpb_processo.setValue(progresso)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())
            self._janela.qpb_processo.setValue(0)
        else:
            self._janela.qpb_processo.setRange(0, len(self._janela.sites))
            self._janela.qte_resultado.clear()
            count = 1
            try:
                for nome_ponto, ip in self._janela.sites.items():
                    for progresso, porta in enumerate(self._janela.portas, 1):
                        retorno = gerar_lista_serial(ip, porta)
                        if retorno:
                            self._janela.qte_resultado.appendPlainText(
                                f"{nome_ponto};{porta};{retorno} \n"
                            )
                    count += 1
                    self._janela.qpb_processo.setValue(count)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())

    def coletar_firmewares(self) -> None:
        """Esse método carrega os dados dos firmewares das câmeras e atualiza a progress bar e o textfield resultado,
        se selecionado o checkbox ele executa em todos os pontos da lista."""
        if not self._janela.qcb_todas_cameras.isChecked():
            self._janela.qpb_processo.setRange(0, len(self._janela.portas[4:]))
            self._janela.qte_resultado.clear()
            ponto = self._janela.pontoComboBox.currentText()
            try:
                for progresso, porta in enumerate(self._janela.portas):
                    retorno = gerar_lista_firmeware(
                        self._janela.sites.get(ponto), porta
                    )
                    if retorno:
                        self._janela.qte_resultado.appendPlainText(
                            f"{ponto};{porta}:{retorno} \n"
                        )
                        self._janela.qpb_processo.setValue(progresso)
            except Exception as e:
                print(e.__str__())
            self._janela.qpb_processo.setValue(0)
        else:
            self._janela.qpb_processo.setRange(0, len(self._janela.sites))
            self._janela.qte_resultado.clear()
            count = 1
            try:
                for nome_ponto, ip in self._janela.sites.items():
                    for progresso, porta in enumerate(self._janela.portas, 1):
                        retorno = gerar_lista_firmeware(ip, porta)
                        if retorno:
                            self._janela.qte_resultado.appendPlainText(
                                f"{nome_ponto};{porta};{retorno} \n"
                            )
                    count += 1
                    self._janela.qpb_processo.setValue(count)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())

    def setar_reset_cameras_contexto(self):
        """Esse método carrega os dados dos serias das câmeras e atualiza a progress bar e o textfield resultado,
        se selecionado o checkbox ele executa em todos os pontos da lista."""
        if not self._janela.qcb_todas_cameras.isChecked():
            self._janela.qpb_processo.setRange(0, len(self._janela._portas[6:]))
            self._janela.qte_resultado.clear()
            ponto = self._janela.pontoComboBox.currentText()
            try:
                for progresso, porta in enumerate(self._janela._portas[6:], 1):
                    retorno = configura_automaintain(
                        self._janela._sites.get(ponto), porta, random.randint(1, 6)
                    )
                    if retorno == 200:
                        self._janela.qte_resultado.appendPlainText(
                            f"{porta} - Reset semanal configurado com sucesso \n"
                        )
                    self._janela.qpb_processo.setValue(progresso)
            except Exception as e:
                print(e.__str__())
            self._janela.qpb_processo.setValue(0)
        else:
            self._janela.qpb_processo.setRange(0, len(self._janela.sites))
            self._janela.qte_resultado.clear()
            count = 1
            try:
                for nome_ponto, ip in self._janela.sites.items():
                    for progresso, porta in enumerate(self._janela.portas, 1):
                        retorno = configura_automaintain(
                            ip, porta, random.randint(1, 6)
                        )
                        if retorno:
                            self._janela.qte_resultado.appendPlainText(
                                f"{nome_ponto};{porta};{retorno} \n"
                            )
                    count += 1
                    self._janela.qpb_processo.setValue(count)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())

    def setar_fontes(self):
        """Esse método configura as fontes para os dois fluxos de stream e atualiza a progress bar e o textfield resultado,
        se selecionado o checkbox ele executa em todos sos pontos da lista."""
        if not self._janela.qcb_todas_cameras.isChecked():
            self._janela.qpb_processo.setRange(0, len(self._janela.portas))
            self._janela.qte_resultado.clear()
            ponto = self._janela.pontoComboBox.currentText()
            try:
                for progresso, porta in enumerate(self._janela.portas):
                    resposta = confiurar_tamanho_fontes(
                        self._janela.sites.get(ponto), porta
                    )
                    self._janela.qte_resultado.appendPlainText(f"{porta};{resposta} \n")
                    self._janela.qpb_processo.setValue(progresso)
            except Exception as e:
                print(f"{e.__str__()} - Trecho do código: {traceback.format_exc()}")

            self._janela.qpb_processo.setValue(0)
        else:
            self._janela.qpb_processo.setRange(0, len(self._janela.sites))
            self._janela.qte_resultado.clear()
            count = 1
            try:
                for nome_ponto, ip in self._janela.sites.items():
                    for progresso, porta in enumerate(self._janela.portas, 1):
                        retorno = confiurar_tamanho_fontes(ip, porta)
                        if retorno:
                            self._janela.qte_resultado.appendPlainText(
                                f"{nome_ponto};{porta};{retorno} \n"
                            )
                    count += 1
                    self._janela.qpb_processo.setValue(count)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())

    def coletar_auto_register(self):
        """Esse método coleta os dados do auto register e atualiza a progress bar e o textfield resultado,
        se selecionado o checkbox ele executa em todos sos pontos da lista."""
        if not self._janela.qcb_todas_cameras.isChecked():
            self._janela.qpb_processo.setRange(0, len(self._janela.portas))
            self._janela.qte_resultado.clear()
            ponto = self._janela.pontoComboBox.currentText()
            try:
                for progresso, porta in enumerate(self._janela.portas):
                    retorno = gerar_lista_portas(self._janela.sites.get(ponto), porta)
                    if retorno:
                        self._janela.qte_resultado.appendPlainText(
                            f"{porta} - {retorno} \n"
                        )
                    self._janela.qpb_processo.setValue(progresso)
            except Exception as e:
                print(e.__str__())
            self._janela.qpb_processo.setValue(0)
        else:
            self._janela.qpb_processo.setRange(0, len(self._janela.sites))
            self._janela.qte_resultado.clear()
            count = 1
            try:
                for nome_ponto, ip in self._janela.sites.items():
                    for progresso, porta in enumerate(self._janela.portas, 1):
                        retorno = gerar_lista_portas(ip, porta)
                        if retorno:
                            self._janela.qte_resultado.appendPlainText(
                                f"{nome_ponto};{porta};{retorno} \n"
                            )
                    count += 1
                    self._janela.qpb_processo.setValue(count)
            except Exception as e:
                print(e.__str__(), traceback.format_exc())
