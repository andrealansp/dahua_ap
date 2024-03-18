import random
import re
import time

import requests
from openpyxl import load_workbook
from requests.auth import HTTPDigestAuth
from requests.exceptions import Timeout

wb = load_workbook(filename="dados.xlsx")
ws = wb.active
listaDadosLido = []
dados_lido_csv = {}


def gerar_lista_portas(ip, porta):
    try:
        urlport = (
            f"http://{ip}:{porta}/cgi-bin/configManager.cgi?action=getConfig&name=DVRIP"
        )
        s = requests.get(
            urlport, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=1
        )
        returneddata = s.text.split("\r\n")
        time.sleep(0.2)
        resposta = returneddata[8]
        print(resposta)
        return re.sub(r"\w{5}.\w{5}.\w*.\w*[\D\w]{3}.\w{4}=", "PORTA:", returneddata[8])
    except Timeout:
        return f"Câmera {porta} Está offline"


def gerar_lista_serial(ip, porta):
    try:
        # ttp://<server>/
        urlport = f"http://{ip}:{porta}/cgi-bin/magicBox.cgi?action=getSerialNo"
        s = requests.get(
            urlport, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=1
        )
        returneddata = s.text.split("\r\n")
        time.sleep(1)
        resposta = returneddata[0]
        print(resposta)
        return returneddata[0]
    except Timeout:
        return f"Câmera {porta} Está offline"


def gerar_lista_modelos(ip, porta):
    try:
        urlport = f"http://{ip}:{porta}/cgi-bin/magicBox.cgi?action=getDeviceType"
        s = requests.get(
            urlport, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=2
        )
        returneddata = s.text.split("\r\n")
        time.sleep(0.2)
        resposta = returneddata[0]
        return re.sub(r"(\w{4}.)(\w{6}.\w{4}.\w{3})", r" \2", returneddata[0])
    except Timeout:
        return f"Câmera {porta} Está offline"


def gerar_lista_firmeware(ip, porta):
    try:
        urlport = f"http://{ip}:{porta}/cgi-bin/magicBox.cgi?action=getSoftwareVersion"
        s = requests.get(
            urlport, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=1
        )
        returneddata = s.text.split("\r\n")
        time.sleep(0.2)
        retorno_firmware = re.sub(
            r"(\w{7}.)(\d.\d{3}.[WG0-9]{7}.\d{1,2}.\w,)(\w{5}.)(\d{4}.\d{2}.\d{2})",
            r"Versão: \2 - Build: \4",
            returneddata[0],
        )
        return retorno_firmware
    except Timeout:
        return f"Câmera {porta} Está offline"
    except requests.exceptions.RetryError:
        return f"Câmera {porta} Está offline"
    except requests.exceptions.ConnectionError:
        return f"Câmera {porta} Está offline"


def listar_enconding_strateg(ip, porta):
    try:
        urlport = f"http://{ip}:{porta}/cgi-bin/configManager.cgi?action=getConfig&name=SmartEncode&name=AICoding"
        s = requests.get(
            urlport, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=1
        )
        returneddata = s.text
        time.sleep(0.5)
        resposta_aicode = re.sub(
            r"(\w{5}).(\w{8,11})\D.\D.(\w{6}.)(\w{3,5})", r"\2 -\4", returneddata
        )
        return resposta_aicode
    except Timeout:
        return f"A Câmera {porta} está offline!"
        pass


def configura_automaintain(ip, porta, opcao):
    opcoes = {
        "1": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 1,
            "AutoMaintain.AutoRebootHour": 0,
            "AutoMaintain.AutoRebootMinute": 0,
        },
        "2": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 2,
            "AutoMaintain.AutoRebootHour": 1,
            "AutoMaintain.AutoRebootMinute": 0,
        },
        "3": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 3,
            "AutoMaintain.AutoRebootHour": 2,
            "AutoMaintain.AutoRebootMinute": 0,
        },
        "4": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 4,
            "AutoMaintain.AutoRebootHour": 3,
            "AutoMaintain.AutoRebootMinute": 0,
        },
        "5": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 5,
            "AutoMaintain.AutoRebootHour": 4,
            "AutoMaintain.AutoRebootMinute": 0,
        },
        "6": {
            "action": "setConfig",
            "AutoMaintain.AutoRebootDay": 6,
            "AutoMaintain.AutoRebootHour": 5,
            "AutoMaintain.AutoRebootMinute": 0,
        },
    }

    try:
        urlport = f"http://{ip}:{porta}/cgi-bin/configManager.cgi"
        s = requests.get(
            urlport,
            params=opcoes[str(opcao)],
            auth=HTTPDigestAuth("admin", "perkons6340"),
            timeout=1,
        )
        returneddata = s.status_code
        time.sleep(0.5)
        return returneddata
    except Timeout:
        print("Camera Offline")
        pass


def configurar_font_size(ip, porta):
    respostas = []
    try:
        urls = [
            f"http://{ip}:{porta}/cgi-bin/configManager.cgi?action=setConfig&VideoWidget[0].FontSize=0",
            f"http://{ip}:{porta}/cgi-bin/configManager.cgi?action=setConfig&VideoWidget[0].FontExtra1=0",
        ]
        for url in urls:
            s = requests.get(
                url, auth=HTTPDigestAuth("admin", "perkons6340"), timeout=1
            )
            returneddata = s.text.split("\r\n")
            time.sleep(0.2)
            respostas.append(returneddata[8])
        return respostas
    except Timeout:
        return f"Câmera {porta} Está offline"


def configurar_reset():
    for row in ws.iter_rows(min_row=2):
        print(
            row[0].value,
            "-",
            configura_automaintain(row[1].value, row[2].value, random.randint(1, 6)),
        )
        print(
            row[0].value,
            "-",
            configura_automaintain(row[1].value, row[3].value, random.randint(1, 6)),
        )


def listar_aicode():
    for row in ws.iter_rows(min_row=2):
        # if ip.find(".", 0, len(ip)) == -1:
        #     row[1].value = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)

        with open("listaraicode.txt", "a", encoding="utf8", newline="") as a:
            a.write(
                f"{row[0].value} - CONT-ENTRADA -  {listar_enconding_strateg(row[1].value, row[2].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONT-SAÍDA -  {listar_enconding_strateg(row[1].value, row[3].value)} \n"
            )


def listar_portas():
    for row in ws.iter_rows(min_row=2):
        # if ip.find(".", 0, len(ip)) == -1:
        #     row[1].value = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)

        with open("ListaPortas.txt", "a", encoding="utf8", newline="") as a:
            a.write(
                f"{row[0].value} - LPR-ENTRADA -  {gerar_lista_portas(row[1].value, row[4].value)} \n"
            )
            a.write(
                f"{row[0].value} - LPR-SAÍDA -  {gerar_lista_portas(row[1].value, row[5].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONT-ENTRADA -  {gerar_lista_portas(row[1].value, row[2].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONT-SAÍDA -  {gerar_lista_portas(row[1].value, row[3].value)} \n"
            )

        print(row[0].value)


def listar_firmeware():
    for row in ws.iter_rows(min_row=2):
        # if ip.find(".", 0, len(ip)) == -1:
        #     row[1].value = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)
        # (\d{1}.\d{3}.\d{7}.\d{2}\W\w)
        # (\d{4}\-?\d{2}\-?\d{2})

        with open("listadefirmeware.txt", "a", encoding="utf8", newline="") as a:
            a.write(
                f"{row[0].value} - LPR-ENTRADA -  {gerar_lista_firmeware(row[1].value, row[4].value)} \n"
            )
            a.write(
                f"{row[0].value} - LPR-SAÍDA -  {gerar_lista_firmeware(row[1].value, row[5].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONT-ENTRADA -  {gerar_lista_firmeware(row[1].value, row[2].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONT-SAÍDA -  {gerar_lista_firmeware(row[1].value, row[3].value)} \n"
            )

        print(row[0].value)


def listar_modelo():
    for row in ws.iter_rows(min_row=2):
        # if ip.find(".", 0, len(ip)) == -1:
        #     row[1].value = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)
        # (\d{1}.\d{3}.\d{7}.\d{2}\W\w)
        # (\d{4}\-?\d{2}\-?\d{2})

        with open("lista_de_modelos.txt", "a", encoding="utf8", newline="") as a:
            a.write(
                f"{row[0].value} - LPR-ENTRADA -  {gerar_lista_modelos(row[1].value, row[4].value)} \n"
            )
            a.write(
                f"{row[0].value} - LPR-SAÍDA -  {gerar_lista_modelos(row[1].value, row[5].value)} \n"
            )

        print(row[0].value)


def listar_serial():
    for row in ws.iter_rows(min_row=2):
        # if ip.find(".", 0, len(ip)) == -1:
        #     row[1].value = re.sub(r'(\d{2})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)
        # (\d{1}.\d{3}.\d{7}.\d{2}\W\w)
        # (\d{4}\-?\d{2}\-?\d{2})

        with open("lista_de_seriais.txt", "a", encoding="utf8", newline="") as a:
            a.write(
                f"{row[0].value} - LPR-ENTRADA -  {gerar_lista_serial(row[1].value, row[4].value)} \n"
            )
            a.write(
                f"{row[0].value} - LPR-SAÍDA -  {gerar_lista_serial(row[1].value, row[5].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONTEXTO ENTRADA -  {gerar_lista_serial(row[1].value, row[3].value)} \n"
            )
            a.write(
                f"{row[0].value} - CONTEXTO SAIDA -  {gerar_lista_serial(row[1].value, row[4].value)} \n"
            )
        print(row[0].value)


def menu(opcao: int = None):
    while opcao != 100:
        print(
            """
        Escolha a função que deseja realizar:
        
        0 - Gerar uma lista com as configurações de auto register:
        1 - Gerar uma lista com as versões de firmware das câmeras:
        2 - Gerar uma lista com as configurações de encoding strategy     
        3 - Gerar uma lista com os modelos de câmeras.    
        4 - Gerar uma lista com os seriais.    
        """
        )
        opcao = int(input("Digite a opção ou 100 para finalizar:"))

        if opcao == 0:
            listar_portas()
        if opcao == 1:
            listar_firmeware()
        if opcao == 2:
            listar_aicode()
        if opcao == 3:
            listar_modelo()
        if opcao == 4:
            listar_serial()


if __name__ == "__main__":
    menu()
