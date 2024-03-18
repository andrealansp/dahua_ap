import sys

from PyQt5.QtWidgets import QApplication, QDialog, QMessageBox
from PyQt5.uic import loadUi
from openpyxl import load_workbook

from classes.dahua import gerar_lista_serial


class Janela(QDialog):
    def __init__(self) -> None:
        super(Janela, self).__init__()
        loadUi("janela.ui", self)
        self.__sites = []
        self.inicializarformulario()

    def inicializarformulario(self) -> None:
        self.carrega_dados()
        self.open()
        self.btn_coleta_serial.clicked.connect(self.carregar_seriais)

    def open(self) -> None:
        self.show()

    def carrega_dados(self):
        wb = load_workbook("sitesNoDss.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=2):
            self.__sites.append((row[0].value, row[1].value))
        for opcao in self.__sites:
            self.pontoComboBox.addItem(f"{opcao[0]} - {opcao[1]}")

    def carregar_seriais(self):
        self.resultado.clear()
        lista_portas = [9050, 9051, 9070, 9071, 9052, 9053, 9072, 9073]
        try:
            for porta in lista_portas:
                ip = self.pontoComboBox.currentText()
                ip = ip.split(" ")
                self.resultado.append(
                    f"{porta} - {gerar_lista_serial(ip[2], porta)} \n"
                )
        except Exception as e:
            print(e.__str__())

        def sair(self) -> None:
            resposta = QMessageBox.information(
                self,
                "Informação",
                "Deseja Realmente sair do PN10?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if resposta == QMessageBox.Yes:
                self.close()


app = QApplication(sys.argv)
janela = Janela()
janela.open()
sys.exit(app.exec_())
